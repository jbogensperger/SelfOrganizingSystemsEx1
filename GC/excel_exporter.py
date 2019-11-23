from openpyxl import Workbook
from openpyxl.worksheet import Worksheet

from usecases.genetic_algorithms.graph_coloring import Coloring
from usecases.genetic_algorithms.tsp import Tour


def get_workbook_and_worksheet() -> (Workbook, Worksheet):
    wb = instantiate_workbook()
    ws = wb.active

    return wb, ws


def instantiate_workbook() -> Workbook:
    """
    :return: Instantiated workbook where experiment results are stored
    """
    wb = Workbook()
    ws = wb.active

    ws.merge_cells('A1:A2')
    ws['A1'] = 'INSTANCE'
    ws.column_dimensions['A'].width = 30
    ws.row_dimensions[1].height = 23
    ws.row_dimensions[2].height = 23

    ws['B2'] = 'NUMBER OF NODES'
    ws.column_dimensions['B'].width = 16

    ws['C2'] = 'NUMBER OF EDGES'
    ws.column_dimensions['C'].width = 16

    return wb


def write_instance_name(ws, value: str, start_index: int, no_of_runs: int):
    """Write the name of the instance in the corresponding column in the sheet"""
    ws.merge_cells('A{}:A{}'.format(start_index, start_index + no_of_runs - 1))
    ws['A{}'.format(start_index)] = value


def write_parameter_data_tsp(ws, param_index: int, param_value: str):
    """
    Stupid index manovrations to have experiment results exported to excel. Just believe it and don't touch it
    :param ws:
    :param param_index:
    :param param_value:
    :return:
    """
    ws.merge_cells('{}1:{}1'.format(chr((param_index * 3) + (param_index - 1) + 65),
                                    chr((param_index * 3) + (param_index - 1) + 3 + 65)))

    ws['{}1'.format(chr((param_index * 3) + (param_index - 1) + 65))] = param_value

    ws['{}2'.format(chr((param_index * 3) + (param_index - 1) + 65))] = 'FINAL COST'
    ws.column_dimensions[chr((param_index * 3) + (param_index - 1) + 65)].width = 16

    ws['{}2'.format(chr((param_index * 3) + (param_index - 1) + 65 + 1))] = 'VERTEX ORDER'
    ws.column_dimensions[chr((param_index * 3) + (param_index - 1) + 65 + 1)].width = 16

    ws['{}2'.format(chr((param_index * 3) + (param_index - 1) + 65 + 2))] = 'IS INFESIBLE'
    ws.column_dimensions[chr((param_index * 3) + (param_index - 1) + 65 + 2)].width = 16

    ws['{}2'.format(chr((param_index * 3) + (param_index - 1) + 65 + 3))] = 'TIME (s)'
    ws.column_dimensions[chr((param_index * 3) + (param_index - 1) + 65 + 3)].width = 16


def write_experiment_results_tsp_ga(ws, param_index: int, row_number: int, tour: Tour, elapsed_time: float):
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65), row_number)] = tour.get_fitness()
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 1), row_number)] = str(tour)
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 2), row_number)] = 'IMPLEMENT ME!'
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 3), row_number)] = float("{:.3f}".format(elapsed_time))


def write_experiment_results_tsp_aco(ws, param_index: int, row_number: int, res, elapsed_time: float):
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65), row_number)] = res[1]
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 1), row_number)] = ' ,'.join(str(x) for x in res[0])
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 2), row_number)] = 'IMPLEMENT ME!'
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 3), row_number)] = float("{:.3f}".format(elapsed_time))


def write_parameter_data_gc(ws, param_index: int, param_value: str):
    """
    Stupid index manovrations to have experiment results exported to excel. Just believe it and don't touch it
    :param ws:
    :param param_index:
    :param param_value:
    :return:
    """
    ws.merge_cells('{}1:{}1'.format(chr((param_index * 3) + (param_index - 1) + 65),
                                    chr((param_index * 3) + (param_index - 1) + 3 + 65)))

    ws['{}1'.format(chr((param_index * 3) + (param_index - 1) + 65))] = param_value

    ws['{}2'.format(chr((param_index * 3) + (param_index - 1) + 65))] = 'INITIAL NO OF COLORS'
    ws.column_dimensions[chr((param_index * 3) + (param_index - 1) + 65)].width = 16

    ws['{}2'.format(chr((param_index * 3) + (param_index - 1) + 65 + 1))] = 'NO OF COLORS'
    ws.column_dimensions[chr((param_index * 3) + (param_index - 1) + 65 + 1)].width = 16

    ws['{}2'.format(chr((param_index * 3) + (param_index - 1) + 65 + 2))] = 'CONFLICTS'
    ws.column_dimensions[chr((param_index * 3) + (param_index - 1) + 65 + 2)].width = 16

    ws['{}2'.format(chr((param_index * 3) + (param_index - 1) + 65 + 3))] = 'TIME (s)'
    ws.column_dimensions[chr((param_index * 3) + (param_index - 1) + 65 + 3)].width = 16


def write_experiment_results_gc_ga(ws, param_index: int, row_number: int, coloring: Coloring, elapsed_time: float):
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65), row_number)] = coloring.initial_number_of_colors
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 1), row_number)] = coloring.get_number_of_colors()
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 2), row_number)] = coloring.get_number_of_color_conflicts()
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 3), row_number)] = float("{:.3f}".format(elapsed_time))


def write_experiment_results_gc_aco(ws, param_index: int, row_number: int, coloring: Coloring, elapsed_time: float):
    # TODO: FIX ME
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65), row_number)] = coloring.initial_number_of_colors
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 1), row_number)] = coloring.get_number_of_colors()
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 2), row_number)] = coloring.get_number_of_color_conflicts()
    ws['{}{}'.format(
        chr((param_index * 3) + (param_index - 1) + 65 + 3), row_number)] = float("{:.3f}".format(elapsed_time))
